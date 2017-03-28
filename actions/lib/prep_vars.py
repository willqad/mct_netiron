# Copyright 2017 Brocade Communications Systems, Inc.
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at

#   http://www.apache.org/licenses/LICENSE-2.0

# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from openpyxl import load_workbook
from more_itertools import unique_everseen
import subprocess
from st2actions.runners.pythonrunner import Action

__all__ = [
    'PrepVars', 'Vrrpe'
]


class ConnError(Exception):
    def __init__(self):
        Exception.__init__(self, 'One or more of your devices is not reachable')


class PrepVars(Action):

    def __init__(self, config = None):
        try:
            self._wb = load_workbook(filename='details.xlsx', data_only=True)
            self._ws = self._wb['Switch Details']

        except IOError:
            print "File does not exist in the right directory /opt/stackstorm/packs/mct_netiron/actions"

        self.ipadd_sw1 = self._ws['C7'].value
        self.ipadd_sw2 = self._ws['D7'].value
        self.user_sw1 = self._ws['C8'].value
        self.user_sw2 = self._ws['D8'].value
        self.password_sw1 = self._ws['C9'].value
        self.password_sw2 = self._ws['D9'].value
        self.icl1_sw1 = self._ws['C10'].value
        self.icl1_sw2 = self._ws['D10'].value
        self.session_vlan = self._ws['C12'].value
        self.session_network = self._ws['C13'].value
        self.session_subnet = self._ws['C14'].value
        if self._ws['C11'].value and self._ws['D11'].value:
            self.icl2_sw1 = self._ws['C11'].value
            self.icl2_sw2 = self._ws['D11'].value
            self.session_list = [self.session_vlan, self.session_subnet, self.session_network, self.icl1_sw2,
                                 self.icl1_sw1, self.icl2_sw1, self.icl2_sw2]
        else:
            self.session_list = [self.session_vlan, self.session_subnet, self.session_network, self.icl1_sw2,
                                 self.icl1_sw1]
        if self._ws['C15'].value:
            self.kalive_vlan=self._ws['C15'].value
            self.kalive_port_sw1=self._ws['C16'].value
            self.kalive_port_sw2=self._ws['D16'].value
        else:
            raise IOError('No Keep-alive vlan is add in the sheet')

    def lags(self):
        # add the lag ports to two lists while making sure that ports for the lag exists on both sides
        _sw1_lag_ports=[]
        _sw2_lag_ports=[]
        for rows in self._ws['C{}:C{}'.format(18,900)]:         # need to find a better solution for this
            for cell in rows:
                if cell.value:
                    _sw1_lag_ports.append(str(cell.value))
        for rows in self._ws['D{}:D{}'.format(18,900)]:         # need to find a better solution for this
            for cell in rows:
                if cell.value:
                    _sw2_lag_ports.append(str(cell.value))
        if len(_sw1_lag_ports) == len(_sw2_lag_ports):
            return _sw1_lag_ports, _sw2_lag_ports
        else:
            raise IOError('There is a port miss match between the two switches')

    def lags_vlans(self):
        _lag_names = []
        _tagged_vlans = []
        _untagged_vlan = []
        _lags_tagged_vlans = {}
        _lags_untagged_vlan = {}
        _sw1_lag_ports = []
        for rows in self._ws['E{}:E{}'.format(18, 900)]:         # need to find a better solution for this
            for cell in rows:
                if cell.value:
                    _lag_names.append(str(cell.value))
        for rows in self._ws['F{}:F{}'.format(18, 900)]:         # need to find a better solution for this
            for cell in rows:
                if cell.value:
                    _tagged_vlans.append(str(cell.value))
        for rows in self._ws['G{}:G{}'.format(18, 900)]:         # need to find a better solution for this
            for cell in rows:
                if cell.value:
                    _untagged_vlan.append(str(cell.value))
        # Combining both lists of tagged vlans and lag names into 1 Dictionary
        _tagged_zip = list(zip(_lag_names,_tagged_vlans))
        for (x, y) in _tagged_zip:
            _lags_tagged_vlans[x] = y

        # Combining both lists of untagged vlan and lag names into 1 Dictionary
        _untagged_zip = list(zip(_lag_names, _untagged_vlan))
        for (x, y) in _untagged_zip:
            _lags_untagged_vlan[x] = y
        for rows in self._ws['C{}:C{}'.format(18,900)]:         # need to find a better solution for this
            for cell in rows:
                if cell.value:
                    _sw1_lag_ports.append(str(cell.value))
        if len(_lag_names) == len(_sw1_lag_ports):
            return _lag_names, _lags_tagged_vlans, _lags_untagged_vlan
        else:
            raise IOError('miss match between tagged vlans and port count')

    # Create a list of non Duplicate Vlans from all ports
    def vlans(self):
        _vlans = []
        nd_vlans = []
        for rows in self._ws['F{}:F{}'.format(18, 900)]:
            for cell in rows:
                if cell.value:
                    try:
                        cell_vlan_list = cell.value.split(',')
                        _vlans.extend(cell_vlan_list)
                        nd_vlans = list(unique_everseen(_vlans))        # Removes the duplicate vlans
                        nd_vlans = filter(None, nd_vlans)
                    except IOError:
                        print("Please make sure that the vlans list are ending with ' , '")
        return nd_vlans

    # create an IP addresses per session ve per device
    def sessionips(self):
        session_ips= {'session_ip1': '', 'session_ip2': ''}
        ip_add = self._ws['C13'].value
        subnet = self._ws['C14'].value
        mask_octets_padded = []
        mask_octets_decimal = subnet.split(".")
        # print mask_octets_decimal
        for octet_index in range(0, len(mask_octets_decimal)):
            # print bin(int(mask_octets_decimal[octet_index]))
            binary_octet = bin(int(mask_octets_decimal[octet_index])).split("b")[1]
            # print binary_octet
            if len(binary_octet) == 8:
                mask_octets_padded.append(binary_octet)
            elif len(binary_octet) < 8:
                binary_octet_padded = binary_octet.zfill(8)
                mask_octets_padded.append(binary_octet_padded)
            # print mask_octets_padded
        decimal_mask = "".join(mask_octets_padded)
        # print decimal_mask
        # Counting host bits in the mask and calculating number of hosts/subnet
        no_of_zeros = decimal_mask.count("0")
        no_of_ones = 32 - no_of_zeros
        # Convert IP to binary string
        ip_octets_padded = []
        ip_octets_decimal = ip_add.split(".")
        for octet_index in range(0, len(ip_octets_decimal)):
            binary_octet = bin(int(ip_octets_decimal[octet_index])).split("b")[1]
            if len(binary_octet) < 8:
                binary_octet_padded = binary_octet.zfill(8)
                ip_octets_padded.append(binary_octet_padded)
            else:
                ip_octets_padded.append(binary_octet)
        # print ip_octets_padded
        binary_ip = "".join(ip_octets_padded)
        # print binary_ip

        # Obtain the network address and broadcast address from the binary strings obtained above
        network_address_binary = binary_ip[:(no_of_ones)] + "0" * no_of_zeros
        # print network_address_binary

        broadcast_address_binary = binary_ip[:(no_of_ones)] + "1" * no_of_zeros
        # print broadcast_address_binary
        net_ip_octets = []
        for octet in range(0, len(network_address_binary), 8):
            net_ip_octet = network_address_binary[octet:octet + 8]
            net_ip_octets.append(net_ip_octet)

        # print net_ip_octets

        net_ip_address = []
        for each_octet in net_ip_octets:
            net_ip_address.append(str(int(each_octet, 2)))

        # print net_ip_address

        bst_ip_octets = []
        for octet in range(0, len(broadcast_address_binary), 8):
            bst_ip_octet = broadcast_address_binary[octet:octet + 8]
            bst_ip_octets.append(bst_ip_octet)

        # print bst_ip_octets

        bst_ip_address = []
        for each_octet in bst_ip_octets:
            bst_ip_address.append(str(int(each_octet, 2)))

        # Generate Session IPs

        for val in range(1,3):
            generated_ip = []
            for indexb, oct_bst in enumerate(bst_ip_address):
                for indexn, oct_net in enumerate(net_ip_address):
                    if indexb == indexn:
                        if oct_bst == oct_net:
                            # Add identical octets to the generated_ip list
                            generated_ip.append(oct_bst)
                        else:
                            generated_ip.append(str((int(oct_net)+val)))
            session_ips['session_ip{0}'.format(val)] = ".".join(generated_ip)
        return session_ips
    # Check the IP address validity

    def ip_valid(self):
        _ip_list = [self.ipadd_sw1, self.ipadd_sw2]
        for ip in _ip_list:
            _ip_broken = ip.split('.')
            if (len(_ip_broken) == 4) and (1 <= int(_ip_broken[0]) <= 223) and (int(_ip_broken[0]) != 127) and (int(_ip_broken[0]) != 169 or int(_ip_broken[1]) != 254) and (0 <= int(_ip_broken[1]) <= 255 and 0 <= int(_ip_broken[2]) <= 255 and 0 <= int(_ip_broken[3]) <= 255):
                print(' IP addresses correct')
            else:
                raise IOError("IP address shouldn't be a localhost or self assigned IP")

    # Pinging The devices to ensure readability

    def ip_reachable(self):
        _ip_list = [self.ipadd_sw1, self.ipadd_sw2]
        for ip in _ip_list:
            # _ping_reply = subprocess.call(['ping', '-c', '2', '-w', '2', '-q', '-n', ip])
            _ping_reply = subprocess.call(['ping', '-c4', ip])  # testing on MAC
            print ip
            if _ping_reply == 0:
                pass
            else:
                raise ConnError()
    # Prepare Netiron Dictionaries

    def mct_devices(self):

        netiron_sw1 = {
            'device_type': 'brocade_netiron',
            'ip': self.ipadd_sw1,
            'username': self.user_sw1,
            'password': self.password_sw1,
            'port': '22'
            }
        netiron_sw2 = {
            'device_type': 'brocade_netiron',
            'ip': self.ipadd_sw2,
            'username': self.user_sw2,
            'password': self.password_sw2,
            'port': '22'
            }
        netiron_device = [netiron_sw1, netiron_sw2]
        return netiron_device


class Vrrpe(Action):
    def __init__(self, config = None):
        try:
            self._wb = load_workbook(filename='details.xlsx', data_only=True)
            self._wv = self._wb['VRRPE']

        except IOError:
            print "File does not exist in the right directory /opt/stackstorm/packs/mct_netiron/actions"

    def vdetails(self):
        ve, subnet, sw1ip, sw2ip, vip = ([] for i in range(5))
        for rows in self._wv['A{}:A{}'.format(4,900)]:         # need to find a better solution for this
            for cell in rows:
                if cell.value:
                    ve.append(str(cell.value))
        for rows in self._wv['B{}:B{}'.format(4,900)]:         # need to find a better solution for this
            for cell in rows:
                if cell.value:
                    subnet.append(str(cell.value))
        for rows in self._wv['C{}:C{}'.format(4,900)]:         # need to find a better solution for this
            for cell in rows:
                if cell.value:
                    sw1ip.append(str(cell.value))
        for rows in self._wv['D{}:D{}'.format(4,900)]:         # need to find a better solution for this
            for cell in rows:
                if cell.value:
                    sw2ip.append(str(cell.value))
        for rows in self._wv['E{}:E{}'.format(4,900)]:         # need to find a better solution for this
            for cell in rows:
                if cell.value:
                    vip.append(str(cell.value))

        vrrpe_details = {
            've':ve,
            'subnet': subnet,
            'sw1ip': sw1ip,
            'sw2ip': sw2ip,
            'vip': vip
        }
        return vrrpe_details
